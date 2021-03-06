import smtplib, ssl
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import openpyxl
from tkinter import *
from tkinter.filedialog import askopenfilename
import os
import random
import copy

seniors_array = []
wb = None
ws = None

main = Tk()
main.title("Senior Assassin")
main.geometry("420x300")

excel_sheet_label = Label(main, text = "File Name:")
excel_sheet_label.place(x = 5, y = 50)
excel_sheet_name_label = Label(main, text = "No File Initialized")
excel_sheet_name_label.place(x = 80, y = 50)
status_label = Label(main, text = "")
status_label.pack(side = TOP)
frame = Frame(main)
frame.place(x = 5, y = 80)
assassin_list = Listbox(frame, width = 43)
assassin_list.pack(side = "left", fill = "y")
scrollbar = Scrollbar(frame, orient="vertical")
scrollbar.pack(side = RIGHT, fill=Y)
scrollbar.config(command = assassin_list.yview)
assassin_list.config(yscrollcommand=scrollbar.set)
assassin_list.insert(END, "There is currently no data")


def close_window():
    main.destroy()

def workbook_setup():
    global seniors_array
    global wb
    global ws
    file_path = askopenfilename()
    extension = os.path.splitext(file_path)[1]
    basename = os.path.basename(file_path)
    if extension ==".xlsx":
        wb = openpyxl.load_workbook(file_path)
        ws = wb.worksheets[0]
        if ws.cell(row = 1, column = 1).value == "Assassin First":
            status_label.config(text="Workbook Successfully Opened")
            excel_sheet_name_label.config(text=basename)
            assassin_list.delete(0, END)
            for row in range(2, ws.max_row):
                current = row - 2
                current_person = []
                for column in range(1, 6):
                    current_person.append(str(ws.cell(row=row, column=column).value))
                print(current_person)
                seniors_array.append(current_person)
                assassin_list.insert(END, str(seniors_array[current][0]) + " " + str(seniors_array[current][1]) + " - " + str(seniors_array[current][3]) + " " + str(seniors_array[current][4]) + " - " + str(seniors_array[current][2]))
        else:
            status_label.config(text="Error: Not a Valid File")
            wb = None
            ws = None
    else:
        status_label.config(text="Error: Not a Valid File")
        wb = None
        ws = None

def email():
    if wb == None or ws == None:
        status_label.config(text="Error: You Must Initiate a File")

    else:
        for current in range(0, len(seniors_array)):

            #To, From, Message
            fromaddr = "teeron.hajebi@gmail.com"
            toaddr = seniors_array[current][2]
            msg = MIMEMultipart()
            msg['From'] = fromaddr
            msg['To'] = toaddr
            msg['Subject'] = "Your Target (Game #2)"

            #Message Body
            body = seniors_array[current][0] + ",\n" + "You need to assassinate " + seniors_array[current][3] + " " + seniors_array[current][4] + ".\nGood luck."
            msg.attach(MIMEText(body, 'plain'))

            #SMTP Settings
            server = smtplib.SMTP('smtp.gmail.com', 587)
            server.starttls()
            server.login(fromaddr, "" ) #Account password
            text = msg.as_string()
            server.sendmail(fromaddr, toaddr, text)
            server.quit()

init_button = Button(main, text="Initiate Workbook", command = workbook_setup)
init_button.place(x = 70, y = 265)

email_button = Button(main, text="Email", command = email)
email_button.place(x = 210, y = 265)

quit_button = Button(main, text="Quit", command = close_window)
quit_button.place(x = 270, y = 265)

main.mainloop()